"""Spreadsheet + CSV writers for benchmark results."""

from __future__ import annotations

import csv
import json
import statistics
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .pricing import PRICING_VERSION, resolve_pricing

TOOL_VERSION = "1.0.0"

RESULT_HEADERS = [
    "Question ID",
    "Question",
    "Provider",
    "Model",
    "Run #",
    "Success",
    "Final Answer",
    "Generated SQL",
    "SQL Success",
    "SQL Error",
    "Query Result",
    "Input Tokens",
    "Output Tokens",
    "Cached Input Tokens",
    "Cache Write Tokens",
    "Reasoning Tokens",
    "Total Tokens",
    "Input Cost USD",
    "Output Cost USD",
    "Cache Cost USD",
    "Estimated Total Cost USD",
    "SQL Gen Tokens In",
    "SQL Gen Tokens Out",
    "Narrate Tokens In",
    "Narrate Tokens Out",
    "LLM Latency ms",
    "SQL Latency ms",
    "Total Latency ms",
    "Retries",
    "Error Kind",
    "Error",
    "Prompt Version",
    "Pricing Version",
    "Pricing Found",
    "Timestamp",
]


def _median(vals: Sequence[float]) -> Optional[float]:
    if not vals:
        return None
    return float(statistics.median(vals))


def _mean(vals: Sequence[float]) -> Optional[float]:
    if not vals:
        return None
    return float(statistics.mean(vals))


def serialize_rows(rows: Any, limit: int = 50) -> str:
    if rows is None:
        return ""
    try:
        clipped = rows[:limit] if isinstance(rows, list) else rows
        return json.dumps(clipped, default=str, ensure_ascii=False)
    except Exception:  # noqa: BLE001
        return str(rows)[:5000]


def row_from_result(
    *,
    question_id: str,
    question: str,
    run_number: int,
    result: Any,
    timestamp: str,
) -> Dict[str, Any]:
    usage = result.total_usage
    cost = result.cost or {}
    sql_call = next((c for c in result.llm_calls if c.stage == "sql_generation"), None)
    narrate_call = next((c for c in result.llm_calls if c.stage == "result_narration"), None)

    return {
        "Question ID": question_id,
        "Question": question,
        "Provider": result.provider,
        "Model": result.model,
        "Run #": run_number,
        "Success": bool(result.success),
        "Final Answer": result.answer or "",
        "Generated SQL": result.sql or "",
        "SQL Success": result.sql_execution_success,
        "SQL Error": result.sql_error or "",
        "Query Result": serialize_rows(result.rows),
        "Input Tokens": usage.input_tokens,
        "Output Tokens": usage.output_tokens,
        "Cached Input Tokens": usage.cached_input_tokens,
        "Cache Write Tokens": usage.cache_write_tokens,
        "Reasoning Tokens": usage.reasoning_tokens,
        "Total Tokens": usage.total_tokens,
        "Input Cost USD": cost.get("estimated_input_cost_usd"),
        "Output Cost USD": cost.get("estimated_output_cost_usd"),
        "Cache Cost USD": cost.get("estimated_cache_cost_usd"),
        "Estimated Total Cost USD": cost.get("estimated_total_cost_usd"),
        "SQL Gen Tokens In": sql_call.usage.input_tokens if sql_call else None,
        "SQL Gen Tokens Out": sql_call.usage.output_tokens if sql_call else None,
        "Narrate Tokens In": narrate_call.usage.input_tokens if narrate_call else None,
        "Narrate Tokens Out": narrate_call.usage.output_tokens if narrate_call else None,
        "LLM Latency ms": result.total_llm_latency_ms,
        "SQL Latency ms": result.sql_execution_ms,
        "Total Latency ms": result.total_latency_ms,
        "Retries": result.retries_total,
        "Error Kind": result.error_kind or "",
        "Error": result.error or "",
        "Prompt Version": result.prompt_version,
        "Pricing Version": cost.get("pricing_version") or PRICING_VERSION,
        "Pricing Found": cost.get("pricing_found"),
        "Timestamp": timestamp,
    }


def summarize_by_model(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Per provider/model summary.

    Average cost per question uses the mean of individual execution costs
    (one row = one model execution). Projected monthly costs use that same
    per-execution average — not (runs × questions) — so a consistency test
    with runs=5 does not inflate production projections.
    """
    groups: Dict[tuple, List[Dict[str, Any]]] = {}
    for r in rows:
        key = (r["Provider"], r["Model"])
        groups.setdefault(key, []).append(r)

    summaries = []
    projections = (100, 500, 1000, 5000, 10000)
    for (provider, model), items in sorted(groups.items()):
        attempted = len(items)
        successes = [i for i in items if i["Success"]]
        failed = attempted - len(successes)
        costs = [i["Estimated Total Cost USD"] for i in items if i["Estimated Total Cost USD"] is not None]
        latencies = [i["Total Latency ms"] for i in items if i["Total Latency ms"] is not None]

        def _sum(field: str) -> Optional[int]:
            vals = [i[field] for i in items if i[field] is not None]
            return sum(vals) if vals else None

        avg_cost = _mean(costs)
        summary = {
            "Provider": provider,
            "Model": model,
            "Questions Attempted (executions)": attempted,
            "Successful Executions": len(successes),
            "Failed Executions": failed,
            "Success Rate": (len(successes) / attempted) if attempted else None,
            "Total Input Tokens": _sum("Input Tokens"),
            "Total Output Tokens": _sum("Output Tokens"),
            "Total Cached Tokens": _sum("Cached Input Tokens"),
            "Total Tokens": _sum("Total Tokens"),
            "Total Estimated Cost (this run)": sum(costs) if costs else None,
            "Average Cost / Execution": avg_cost,
            "Median Cost / Execution": _median(costs),
            "Min Cost / Execution": min(costs) if costs else None,
            "Max Cost / Execution": max(costs) if costs else None,
            "Average Latency ms": _mean(latencies),
            "Median Latency ms": _median(latencies),
            "Pricing Found": resolve_pricing(model) is not None,
        }
        for n in projections:
            summary[f"Projected Cost — {n:,} Questions"] = (
                None if avg_cost is None else avg_cost * n
            )
        summaries.append(summary)
    return summaries


def side_by_side_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """One row per question_id × run, with OpenAI / Claude / Gemini columns."""
    keyed: Dict[tuple, Dict[str, Dict[str, Any]]] = {}
    for r in rows:
        key = (r["Question ID"], r["Run #"], r["Question"])
        keyed.setdefault(key, {})[r["Provider"]] = r

    out = []
    for (qid, run, question), by_provider in sorted(keyed.items()):
        openai = by_provider.get("openai", {})
        anthropic = by_provider.get("anthropic", {})
        gemini = by_provider.get("gemini", {})
        out.append(
            {
                "Question ID": qid,
                "Question": question,
                "Run #": run,
                "OpenAI Model": openai.get("Model", ""),
                "OpenAI Final Answer": openai.get("Final Answer", ""),
                "OpenAI SQL": openai.get("Generated SQL", ""),
                "OpenAI Cost": openai.get("Estimated Total Cost USD"),
                "OpenAI Latency ms": openai.get("Total Latency ms"),
                "OpenAI Success": openai.get("Success"),
                "Claude Model": anthropic.get("Model", ""),
                "Claude Final Answer": anthropic.get("Final Answer", ""),
                "Claude SQL": anthropic.get("Generated SQL", ""),
                "Claude Cost": anthropic.get("Estimated Total Cost USD"),
                "Claude Latency ms": anthropic.get("Total Latency ms"),
                "Claude Success": anthropic.get("Success"),
                "Gemini Model": gemini.get("Model", ""),
                "Gemini Final Answer": gemini.get("Final Answer", ""),
                "Gemini SQL": gemini.get("Generated SQL", ""),
                "Gemini Cost": gemini.get("Estimated Total Cost USD"),
                "Gemini Latency ms": gemini.get("Total Latency ms"),
                "Gemini Success": gemini.get("Success"),
                "Preferred Answer": "",
                "OpenAI Score": "",
                "Claude Score": "",
                "Gemini Score": "",
                "Reviewer Notes": "",
            }
        )
    return out


def _style_sheet(ws, money_cols: Optional[List[int]] = None, wrap_cols: Optional[List[int]] = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    money_cols = money_cols or []
    wrap_cols = wrap_cols or []
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), 60))
            if col_idx in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"
        ws.column_dimensions[letter].width = max(12, min(max_len + 2, 48))


def write_workbook(
    *,
    path: Path,
    result_rows: List[Dict[str, Any]],
    metadata: Dict[str, Any],
) -> None:
    wb = Workbook()

    # Sheet 1: Results
    ws1 = wb.active
    ws1.title = "Results"
    ws1.append(RESULT_HEADERS)
    for r in result_rows:
        ws1.append([r.get(h) for h in RESULT_HEADERS])
    money_idx = [
        RESULT_HEADERS.index(h) + 1
        for h in (
            "Input Cost USD",
            "Output Cost USD",
            "Cache Cost USD",
            "Estimated Total Cost USD",
        )
    ]
    wrap_idx = [
        RESULT_HEADERS.index(h) + 1
        for h in ("Question", "Final Answer", "Generated SQL", "Query Result", "Error")
    ]
    _style_sheet(ws1, money_cols=money_idx, wrap_cols=wrap_idx)

    # Sheet 2: Model Summary
    ws2 = wb.create_sheet("Model Summary")
    summaries = summarize_by_model(result_rows)
    if summaries:
        headers = list(summaries[0].keys())
        ws2.append(headers)
        for s in summaries:
            ws2.append([s.get(h) for h in headers])
        money_cols = [
            i + 1
            for i, h in enumerate(headers)
            if "Cost" in h or h.startswith("Projected")
        ]
        _style_sheet(ws2, money_cols=money_cols)
    else:
        ws2.append(["No results"])

    # Sheet 3: Side-by-Side
    ws3 = wb.create_sheet("Side-by-Side")
    sbs = side_by_side_rows(result_rows)
    if sbs:
        headers = list(sbs[0].keys())
        ws3.append(headers)
        for row in sbs:
            ws3.append([row.get(h) for h in headers])
        money_cols = [headers.index(h) + 1 for h in ("OpenAI Cost", "Claude Cost", "Gemini Cost") if h in headers]
        wrap_cols = [
            headers.index(h) + 1
            for h in (
                "Question",
                "OpenAI Final Answer",
                "Claude Final Answer",
                "Gemini Final Answer",
                "OpenAI SQL",
                "Claude SQL",
                "Gemini SQL",
                "Reviewer Notes",
            )
            if h in headers
        ]
        _style_sheet(ws3, money_cols=money_cols, wrap_cols=wrap_cols)
    else:
        ws3.append(["No results"])

    # Sheet 4: Metadata
    ws4 = wb.create_sheet("Metadata")
    ws4.append(["Key", "Value"])
    for k, v in metadata.items():
        if isinstance(v, (dict, list)):
            v = json.dumps(v, default=str)
        ws4.append([k, v])
    _style_sheet(ws4)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_csv(path: Path, result_rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=RESULT_HEADERS)
        writer.writeheader()
        for r in result_rows:
            writer.writerow({h: r.get(h) for h in RESULT_HEADERS})


def timestamp_slug(dt: Optional[datetime] = None) -> str:
    dt = dt or datetime.now(timezone.utc).astimezone()
    return dt.strftime("%Y-%m-%d-%H%M")
